"""Excel loader for plant knowledge ingestion — TICKET-046.

Loads the updated Excel source, validates required headers, and strips
source-management (admin) columns before returning rows for ingestion.
Ignored columns are excluded so they cannot affect source_row_hash.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

DEFAULT_SHEET_NAME = "전체식물_분류정보"

REQUIRED_HEADERS: frozenset[str] = frozenset(
    [
        "한국명",
        "학명",
        "농사로_매칭",
        "농사로ID",
        "과명",
        "원산지",
        "기능성설명",
        "용도",
        "생육형태",
        "생장속도",
        "생육온도",
        "겨울최저온도",
        "습도",
        "광요구도",
        "관리수준",
        "관리요구도",
        "토양",
        "비료",
        "물주기_봄",
        "물주기_여름",
        "물주기_가을",
        "물주기_겨울",
        "병충해",
        "병충해관리",
        "독성",
        "냄새",
        "잎형태",
        "잎색",
        "꽃색",
        "꽃피는계절",
        "성장높이(cm)",
        "성장너비(cm)",
        "번식방법",
        "배치장소",
    ]
)

IGNORED_COLUMNS: frozenset[str] = frozenset(
    [
        "번호",
        "수정이유_분류정보",
        "참고링크_분류정보",
        "확인필요_항목",
    ]
)


class ExcelHeaderError(ValueError):
    """Raised when the workbook is missing required headers or the target sheet."""


def load_workbook_rows(
    path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> tuple[list[str], list[tuple[Any, ...]]]:
    """Open *path*, validate headers, strip ignored columns, return (headers, data_rows).

    Returns ([], []) when the sheet has no rows at all.
    Raises ExcelHeaderError if the sheet is missing or required headers are absent.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ExcelHeaderError(
                f"Sheet '{sheet_name}' not found in {path.name}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not all_rows:
        return [], []

    raw_headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]

    header_set = set(raw_headers)
    missing = REQUIRED_HEADERS - header_set
    if missing:
        raise ExcelHeaderError(
            f"Missing required headers in '{sheet_name}': {sorted(missing)}"
        )

    keep_indices = [i for i, h in enumerate(raw_headers) if h not in IGNORED_COLUMNS]
    headers = [raw_headers[i] for i in keep_indices]

    data_rows: list[tuple[Any, ...]] = [
        tuple(row[i] if i < len(row) else None for i in keep_indices)
        for row in all_rows[1:]
    ]

    return headers, data_rows
