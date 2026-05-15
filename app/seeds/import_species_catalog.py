"""Excel species catalog importer — TICKET-060A0.

Reads PLANT_KNOWLEDGE_EXCEL_PATH (default: data/전체식물_분류정보_v1_updated_7_2.xlsx)
and upserts rows into species_profiles with stable UUIDs and catalog provenance.

Usage:
    python -m app.seeds.import_species_catalog
    python -m app.seeds.import_species_catalog --dry-run
    python -m app.seeds.import_species_catalog --path /alt/path/file.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.dialects.postgresql import insert

from app.core.config import settings
from app.db.session import AsyncSessionLocal
from app.models.species_profile import SpeciesProfile

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_CATALOG_NS = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")
SOURCE_FILE = "전체식물_분류정보_v1_updated_7_2.xlsx"
SOURCE_VERSION = "v1_updated_7_2"

# Optional columns stored verbatim in metadata_json
_META_COLS = [
    "과명", "기능성설명", "용도", "원산지",
    "생육형태", "생장속도", "생육온도", "겨울최저온도",
    "습도", "광요구도", "관리수준", "관리요구도",
    "토양", "비료",
    "물주기_봄", "물주기_여름", "물주기_가을", "물주기_겨울",
    "병충해", "병충해관리", "독성", "냄새",
    "잎형태", "잎색", "꽃색", "꽃피는계절",
    "성장높이(cm)", "성장너비(cm)", "번식방법", "배치장소",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(val: Any) -> str | None:
    """Coerce a cell value to a stripped string, or None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def catalog_uuid(korean_name: str, scientific_name: str | None) -> uuid.UUID:
    """Return a stable UUID5 keyed on the normalized Korean + scientific name."""
    ko = (korean_name or "").strip().lower()
    sci = (scientific_name or "").strip().lower()
    return uuid.uuid5(_CATALOG_NS, f"{ko}::{sci}")


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def load_rows(excel_path: str | Path) -> list[dict]:
    """Parse the Excel file and return a list of row dicts keyed by header."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel catalog not found: {path.resolve()}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    raw_headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    rows: list[dict] = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row_vals)))

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Row → DB record
# ---------------------------------------------------------------------------

def build_record(row: dict, now: datetime) -> dict | None:
    """Convert one Excel row to a species_profiles upsert dict.

    Returns None when Korean name is absent (row must be skipped).
    """
    korean_name = _str(row.get("한국명"))
    if not korean_name:
        return None

    scientific_name = _str(row.get("학명"))
    species_id = catalog_uuid(korean_name, scientific_name)

    care_level = _str(row.get("관리수준")) or _str(row.get("관리요구도"))

    meta: dict[str, Any] = {
        "source": SOURCE_FILE,
        "source_version": SOURCE_VERSION,
        "catalog_allowed": True,
        "aliases": [],
    }
    for col in _META_COLS:
        val = _str(row.get(col))
        if val is not None:
            meta[col] = val

    return {
        "id": species_id,
        "korean_name": korean_name,
        "scientific_name": scientific_name,
        "common_name": None,
        "care_level": care_level,
        # Numeric range fields: raw text in metadata_json; numeric parsing is
        # deferred to a later enrichment ticket.
        "water_min_pct": None,
        "water_max_pct": None,
        "light_min_lux": None,
        "light_max_lux": None,
        "humidity_min_pct": None,
        "humidity_max_pct": None,
        "temperature_min_c": None,
        "temperature_max_c": None,
        "metadata_json": meta,
        "created_at": now,
        "updated_at": now,
    }


# ---------------------------------------------------------------------------
# Import orchestrator
# ---------------------------------------------------------------------------

async def import_catalog(excel_path: str | Path, dry_run: bool = False) -> dict:
    """Parse the Excel file and upsert rows into species_profiles.

    Returns a summary dict with upserted/skipped counts.
    """
    rows = load_rows(excel_path)
    now = datetime.now(UTC)

    records: list[dict] = []
    skipped = 0
    for row in rows:
        rec = build_record(row, now)
        if rec is None:
            skipped += 1
        else:
            records.append(rec)

    if dry_run:
        return {"mode": "dry_run", "would_upsert": len(records), "skipped": skipped}

    if not records:
        return {"upserted": 0, "skipped": skipped}

    async with AsyncSessionLocal() as session:
        async with session.begin():
            stmt = insert(SpeciesProfile).values(records)
            stmt = stmt.on_conflict_do_update(
                index_elements=["id"],
                set_={
                    "korean_name": stmt.excluded.korean_name,
                    "scientific_name": stmt.excluded.scientific_name,
                    "care_level": stmt.excluded.care_level,
                    "metadata_json": stmt.excluded.metadata_json,
                    "updated_at": stmt.excluded.updated_at,
                },
            )
            await session.execute(stmt)

    return {"upserted": len(records), "skipped": skipped}


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import Excel species catalog into species_profiles."
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse only; no DB writes.")
    parser.add_argument("--path", default=None, help="Override Excel file path.")
    args = parser.parse_args()

    excel_path = args.path or settings.PLANT_KNOWLEDGE_EXCEL_PATH
    result = asyncio.run(import_catalog(excel_path, dry_run=args.dry_run))
    print(result)


if __name__ == "__main__":
    main()
