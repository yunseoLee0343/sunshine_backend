"""TICKET-046 — ExcelLoader unit tests."""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from app.ingestion.excel_loader import (
    DEFAULT_SHEET_NAME,
    IGNORED_COLUMNS,
    REQUIRED_HEADERS,
    ExcelHeaderError,
    load_workbook_rows,
)

_VALID_HEADERS = sorted(REQUIRED_HEADERS)


def _write_xlsx(
    headers: list[str],
    rows: list[list[Any]],
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def test_load_valid_file_returns_rows() -> None:
    path = _write_xlsx(_VALID_HEADERS, [["v"] * len(_VALID_HEADERS)])
    headers, rows = load_workbook_rows(path)
    assert len(rows) == 1
    assert len(headers) == len(_VALID_HEADERS)


def test_load_empty_sheet_returns_empty() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DEFAULT_SHEET_NAME
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    headers, rows = load_workbook_rows(Path(tmp.name))
    assert headers == []
    assert rows == []


def test_ignored_columns_filtered_from_headers() -> None:
    headers_with_ignored = _VALID_HEADERS + sorted(IGNORED_COLUMNS)
    path = _write_xlsx(headers_with_ignored, [["v"] * len(headers_with_ignored)])
    headers, _ = load_workbook_rows(path)
    for col in IGNORED_COLUMNS:
        assert col not in headers


def test_ignored_columns_filtered_from_row_values() -> None:
    """Row tuples must not contain values from ignored-column positions."""
    ignored_sorted = sorted(IGNORED_COLUMNS)
    headers_with_ignored = _VALID_HEADERS + ignored_sorted
    base_vals = ["v"] * len(_VALID_HEADERS)
    ignored_vals = ["IGNORED_SENTINEL"] * len(ignored_sorted)
    path = _write_xlsx(headers_with_ignored, [base_vals + ignored_vals])
    _, rows = load_workbook_rows(path)
    assert len(rows) == 1
    assert all(v != "IGNORED_SENTINEL" for v in rows[0])


def test_row_length_matches_filtered_header_count() -> None:
    headers_with_ignored = _VALID_HEADERS + sorted(IGNORED_COLUMNS)
    path = _write_xlsx(headers_with_ignored, [["v"] * len(headers_with_ignored)])
    headers, rows = load_workbook_rows(path)
    assert len(rows[0]) == len(headers)


def test_missing_required_header_raises() -> None:
    incomplete = _VALID_HEADERS[1:]
    path = _write_xlsx(incomplete, [])
    with pytest.raises(ExcelHeaderError, match="Missing required headers"):
        load_workbook_rows(path)


def test_wrong_sheet_name_raises() -> None:
    path = _write_xlsx(_VALID_HEADERS, [], sheet_name="wrong_sheet")
    with pytest.raises(ExcelHeaderError, match="Sheet"):
        load_workbook_rows(path)


def test_new_filename_sheet_name_constant() -> None:
    assert DEFAULT_SHEET_NAME == "전체식물_분류정보"


def test_required_headers_contains_new_columns() -> None:
    new_cols = (
        "물주기_봄", "물주기_여름", "물주기_가을", "물주기_겨울",
        "꽃색", "꽃피는계절", "성장높이(cm)", "배치장소",
    )
    for col in new_cols:
        assert col in REQUIRED_HEADERS, f"{col!r} not in REQUIRED_HEADERS"


def test_ignored_columns_set_is_correct() -> None:
    assert IGNORED_COLUMNS == frozenset(["번호", "수정이유_분류정보", "참고링크_분류정보", "확인필요_항목"])
