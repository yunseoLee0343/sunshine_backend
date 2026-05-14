"""TICKET-046 — Idempotency tests for plant knowledge ingestion."""

from __future__ import annotations

import tempfile
import uuid
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest

from app.ingestion.excel_loader import (
    DEFAULT_SHEET_NAME,
    IGNORED_COLUMNS,
    REQUIRED_HEADERS,
    load_workbook_rows,
)
from app.services.plant_knowledge_ingest_service import (
    _build_col_index,
    _row_hash,
)

_VALID_HEADERS = sorted(REQUIRED_HEADERS)
_IGNORED_SORTED = sorted(IGNORED_COLUMNS)


def _write_xlsx_with_ignored(base_row: list, ignored_vals: list, sheet_name: str = DEFAULT_SHEET_NAME) -> Path:
    headers = _VALID_HEADERS + _IGNORED_SORTED
    row = base_row + ignored_vals
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def test_hash_unchanged_for_identical_rows() -> None:
    row = ("K001", "몬스테라", "Monstera deliciosa")
    assert _row_hash(list(row)) == _row_hash(list(row))


def test_hash_changes_when_care_field_changes() -> None:
    row1 = ("K001", "몬스테라", "Monstera deliciosa", "18~24℃")
    row2 = ("K001", "몬스테라", "Monstera deliciosa", "20~26℃")
    assert _row_hash(list(row1)) != _row_hash(list(row2))


def test_ignored_columns_do_not_affect_row_hash() -> None:
    """Rows differing only in ignored columns produce identical filtered rows and hashes."""
    base = ["content_val"] * len(_VALID_HEADERS)

    path1 = _write_xlsx_with_ignored(base, ["admin1", "link1", "reason1", "check1"])
    path2 = _write_xlsx_with_ignored(base, ["admin2", "link2", "reason2", "check2"])

    _, rows1 = load_workbook_rows(path1)
    _, rows2 = load_workbook_rows(path2)

    assert rows1[0] == rows2[0]
    assert _row_hash(list(rows1[0])) == _row_hash(list(rows2[0]))


def test_ignored_columns_present_would_change_raw_hash() -> None:
    """Confirms ignored columns WOULD affect hash if not stripped — proving filtering matters."""
    row_with_admin1 = ["v"] * len(_VALID_HEADERS) + ["admin_note_A"] * len(_IGNORED_SORTED)
    row_with_admin2 = ["v"] * len(_VALID_HEADERS) + ["admin_note_B"] * len(_IGNORED_SORTED)
    assert _row_hash(row_with_admin1) != _row_hash(row_with_admin2)


@pytest.mark.asyncio
async def test_re_ingest_same_row_is_ignored() -> None:
    from app.models.plant_knowledge_entry import PlantKnowledgeEntry
    from app.models.plant_knowledge_source import PlantKnowledgeSource
    from app.services.plant_knowledge_ingest_service import PlantKnowledgeIngestService

    session = AsyncMock()
    svc = PlantKnowledgeIngestService(session)

    raw_row = ("K001", "몬스테라", "Monstera deliciosa")
    existing_hash = _row_hash(list(raw_row))

    fake_entry = MagicMock(spec=PlantKnowledgeEntry)
    fake_entry.id = uuid.uuid4()
    fake_source = MagicMock(spec=PlantKnowledgeSource)
    fake_source.source_row_hash = existing_hash

    svc._find_entry = AsyncMock(return_value=fake_entry)
    svc._find_latest_source = AsyncMock(return_value=fake_source)

    col_index = _build_col_index(["농사로ID", "한국명", "학명"])
    status = await svc._process_row(
        raw_row=raw_row,
        col_index=col_index,
        row_number=2,
        source_file="test.xlsx",
    )
    assert status == "ignored"


@pytest.mark.asyncio
async def test_re_ingest_changed_row_is_updated() -> None:
    from app.models.plant_knowledge_entry import PlantKnowledgeEntry
    from app.models.plant_knowledge_source import PlantKnowledgeSource
    from app.services.plant_knowledge_ingest_service import PlantKnowledgeIngestService

    session = AsyncMock()
    svc = PlantKnowledgeIngestService(session)

    raw_row = ("K001", "몬스테라", "Monstera deliciosa")
    fake_entry = MagicMock(spec=PlantKnowledgeEntry)
    fake_entry.id = uuid.uuid4()
    fake_source = MagicMock(spec=PlantKnowledgeSource)
    fake_source.source_row_hash = "stale_hash_that_differs"

    svc._find_entry = AsyncMock(return_value=fake_entry)
    svc._find_latest_source = AsyncMock(return_value=fake_source)
    svc._update_children = AsyncMock()
    svc._add_source = AsyncMock()

    col_index = _build_col_index(["농사로ID", "한국명", "학명"])
    status = await svc._process_row(
        raw_row=raw_row,
        col_index=col_index,
        row_number=2,
        source_file="test.xlsx",
    )
    assert status == "updated"
